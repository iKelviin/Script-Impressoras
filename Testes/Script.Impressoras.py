import time
from selenium import webdriver 
from selenium.webdriver.common.by import By 
import openpyxl
import win32com.client as win32
import pandas as pd
from optparse import Option
import requests
import time
import numpy as np
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.message import EmailMessage
from email.mime.base import MIMEBase 
from email import encoders 




print('Iniciando Sistema...')
class Relatorio_Impressoras:
    def Inicio(self):
        self.offline = 0
        self.MFP_432()
        self.Salvar_Dados_MFP_432()
        self.MFP_E52645()
        self.Salvar_Dados_MFP_E52645()
        self.MFP_M479fdw()
        self.Salvar_Dados_MFP_M479fdw()
        self.Enviar_Email()
      
    def MFP_432 (self):
        print('Modelo MFP 432...')
        self.lista_dos_toner = []
        self.lista_da_imagem = []
        self.nome_impressoras = ['Juridico', 'Segurança do Trabalho',
                                'Ambulatorio Dra Renata','TI',
                                'RH','Pré-Impressão',
                                'Ambulatório','Comercial',
                                'Papel e Tinta','Expedição',
                                'Manutenção','Portaria',
                                'Produção','Impressão Digital',
                                'Recepção']
        self.modelo_mfp_432 = [
        'https://10.10.4.150/sws/index.html', #Juridico
        'https://10.10.4.173/sws/index.html', #Segurança do Trabalho
        'https://10.10.4.177/sws/index.html', #Ambulatorio Dra Renata
        'https://10.10.4.174/sws/index.html', #TI
        'https://10.10.4.157/sws/index.html', #RH
        'https://10.10.4.151/sws/index.html', #Pré-Impressão
        'https://10.10.4.155/sws/index.html', #Ambulatório
        'https://10.10.4.162/sws/index.html', #Comercial
        'https://10.10.4.164/sws/index.html', #Papel e Tinta
        'https://10.10.4.163/sws/index.html', #Expedição
        'https://10.10.4.158/sws/index.html', #Manutenção
        'https://10.10.4.160/sws/index.html', #Portaria
        'https://10.10.4.156/sws/index.html', #Produção
        'https://10.10.4.169/sws/index.html', #Impressão Digital
        'https://10.10.4.182/sws/index.html', #Recepção
        ]

        print('Abrindo Navegador')
        self.navegador = webdriver.Chrome()
        time.sleep(3)

        for self.i in range(15):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_mfp_432[self.i])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                time.sleep(35)
                print("Coletando Dados...")
                self.toner = self.navegador.find_element(By.XPATH,'//*[@id="ext-gen300"]/div/table/tbody/tr/td[2]/div/div/div[2]')
                self.lista_dos_toner.append(self.toner.text)
                time.sleep(2)
                print('Armazenando dados do Toner...')
                self.imagem = self.navegador.find_element(By.XPATH,'//*[@id="ext-gen344"]/div/table/tbody/tr/td[2]/div/div/div[2]')
                self.lista_da_imagem.append(self.imagem.text)
                time.sleep(2)
                print('Armazenando dados da Imagem...')    
                print(self.lista_dos_toner, self.lista_da_imagem)  
                time.sleep(3)          
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel('CONTROLE_DE_IMPRESSORAS.xlsx',sheet_name='MFP 432')
                self.lista_dos_toner.append(self.planilha.iloc[self.i][f"% de Toner"])
                self.lista_da_imagem.append(self.planilha.iloc[self.i]["Uni.Imagem"])   
                print(self.lista_dos_toner, self.lista_da_imagem) 

        print("Impressoras MFP 432 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()

    def Salvar_Dados_MFP_432 (self):
        print('Carregando Planilha...')
        self.index = 2
        self.book= openpyxl.load_workbook('CONTROLE_DE_IMPRESSORAS.xlsx')
        self.planilha = self.book['MFP 432']
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.lista_dos_toner, self.lista_da_imagem):
                self.planilha.cell(column=3, row=self.index, value=self.nome)
                self.planilha.cell(column=4, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save('CONTROLE_DE_IMPRESSORAS.xlsx')        
        print("Planilha Salva com Sucesso")

    def MFP_E52645(self):
        print('Modelo MFP E52645...')
        self.cartucho_preto = []
        self.nome_impressoras = ['ALMOXARIFADO','QUALIDADE','ATENDIMENTO','FINANCEIRO','COMPRAS']
        self.kit_aliment_documentos = []
        self.modelo_e52645 = [
        'https://10.10.4.152/hp/device/DeviceStatus/Index', #Almoxarifado
        'https://10.10.4.153/', #Qualidade
        'https://10.10.4.154/', #Atendimento PB
        'https://10.10.4.171/', #Financeiro
        'https://10.10.4.159/', #Compras
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome()
        time.sleep(3)
        for self.l in range(5):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_e52645[self.l])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                time.sleep(35)
                print('Coletando dados...')
                self.cartucho = self.navegador.find_element(By.XPATH,'//*[@id="SupplyPLR0"]')
                print('Armazenando dados do Cartucho...')
                self.cartuchop = self.cartucho.text.split('*')
                self.cartucho_preto.append(self.cartuchop[0])
                self.kit = self.navegador.find_element(By.XPATH,'//*[@id="SupplyPLR1"]')
                print('Armazenado dados Do kit...')
                self.kitA = self.kit.text.split('*')
                self.kit_aliment_documentos.append(self.kitA[0])
                time.sleep(3)
                print(self.cartucho_preto, self.kit_aliment_documentos)
                time.sleep(3)
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel('CONTROLE_DE_IMPRESSORAS.xlsx',sheet_name='MFP E52645')

                self.cartucho_preto.append(self.planilha.iloc[self.i]["Toner Preto"])
                self.kit_aliment_documentos.append(self.planilha.iloc[self.i]["Kit Aliment. Doc."])   
                print(self.cartucho_preto, self.kit_aliment_documentos) 

        print("Impressoras MFP E52645 Offline: ",self.offline)
        self.offline = 0
        self.navegador.quit()

    def Salvar_Dados_MFP_E52645 (self):
        print('Carregando Planilha...')
        self.index = 2
        self.book= openpyxl.load_workbook('CONTROLE_DE_IMPRESSORAS.xlsx')
        self.planilha = self.book['MFP E52645']
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.cartucho_preto, self.kit_aliment_documentos):
                self.planilha.cell(column=3, row=self.index, value=self.nome)
                self.planilha.cell(column=4, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save('CONTROLE_DE_IMPRESSORAS.xlsx')        
        print("Planilha Salva com Sucesso")

    def MFP_M479fdw(self):
        print('Modelo MFP M479fdw...')
        self.cartucho_preto = []
        self.cartucho_cyan = []
        self.cartucho_magenta = []
        self.cartucho_yellow = []
        self.modelo_M479fdw = [
        'https://10.10.4.178/#hId-pgConsumables',
        'https://10.10.4.167/#hId-pgConsumables',
        'https://10.10.4.168/#hId-pgConsumables',
        'https://10.10.4.175/#hId-pgConsumables',
        'https://10.10.4.172/#hId-pgConsumables',
        'https://10.10.4.170/#hId-pgConsumables'
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome()
        time.sleep(3)

        for self.j in range(6):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_M479fdw[self.j])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                time.sleep(35)
                print('Coletando dados...')
                self.preto = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[2]')
                self.cyan = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[3]')
                self.magenta = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[4]')
                self.yellow = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[5]')
                print('Armazenando dados dos Cartuchos...')
                self.cartucho_preto.append(self.preto.text)
                self.cartucho_cyan.append(self.cyan.text)
                self.cartucho_magenta.append(self.magenta.text)
                self.cartucho_yellow.append(self.yellow.text)
                time.sleep(3)
                print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
                time.sleep(3)
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel('CONTROLE_DE_IMPRESSORAS.xlsx',sheet_name='MFP M479fdw')

                self.cartucho_preto.append(self.planilha.iloc[self.i]["Preto"])
                self.cartucho_cyan.append(self.planilha.iloc[self.i]["Ciano"])   
                self.cartucho_magenta.append(self.planilha.iloc[self.i]["Magenta"])
                self.cartucho_yellow.append(self.planilha.iloc[self.i]["Amarelo"])
                print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow) 

        print("Impressoras MFP M479fdw Offline: ",self.offline)
        self.offline = 0                
        print('Fechando Navegador...')
        self.navegador.quit()    

    def Salvar_Dados_MFP_M479fdw (self):
        print('Carregando Planilha...')
        self.index = 2
        self.book= openpyxl.load_workbook('CONTROLE_DE_IMPRESSORAS.xlsx')
        self.planilha = self.book['MFP M479fdw']
        for self.i in range(1):
            for self.colorblack, self.colorcyan, self.colormagenta, self.coloryellow in zip(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow):
                self.planilha.cell(column=3, row=self.index, value=self.colorblack)
                self.planilha.cell(column=4, row=self.index, value=self.colorcyan)
                self.planilha.cell(column=5, row=self.index, value=self.colormagenta)
                self.planilha.cell(column=6, row=self.index, value=self.coloryellow)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save('CONTROLE_DE_IMPRESSORAS.xlsx')        
        print("Planilha Salva com Sucesso")

    def Enviar_Email(self):
        self.tipos_colunas = {'Modelo' : str,'Estoque': str}
        self.lista_colunas = ['Modelo','Cor','Estoque']

        self.tabela_simpress = pd.read_excel(r'\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx', sheet_name='Python')

        pd.set_option('display.precision',0)

        self.Toner_MFP432 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "432FDN","Estoque"]

        self.Toner_MFPE52645 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "E52645","Estoque"]

        self.Toner_MFP7E77830 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "HP 7E77830","Estoque"]

        self.Toner_MFP479 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "M479","Estoque"]

        print('Enviando E-mail...')
        time.sleep(3)
        self.tabela_e52645 = pd.read_excel(r'C:\Users\kelvin.rocha\Desktop\Area de Trabalho\Kelvin\Python\Automação Python\Breatifulsoap\CONTROLE_DE_IMPRESSORAS.xlsx', sheet_name='MFP E52645')
        self.tabela_mfp432 = pd.read_excel(r'C:\Users\kelvin.rocha\Desktop\Area de Trabalho\Kelvin\Python\Automação Python\Breatifulsoap\CONTROLE_DE_IMPRESSORAS.xlsx', sheet_name='MFP 432')
        self.tabela_mfp479 = pd.read_excel(r'C:\Users\kelvin.rocha\Desktop\Area de Trabalho\Kelvin\Python\Automação Python\Breatifulsoap\CONTROLE_DE_IMPRESSORAS.xlsx', sheet_name='MFP M479fdw')        
        
        self.relatorio1 = self.tabela_mfp432.loc[:,"Local":"Uni.Imagem"].style.set_caption('Consumo de Toner HP Laser MFP 432').set_table_styles([
             { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
              'padding-right': '20px'}).hide_index()
      
        self.relatorio2 = self.tabela_e52645.loc[:,"Local":"Kit Aliment. Doc."].style.set_caption('Consumo de Toner HP LaserJet MFP E52645').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.relatorio3 = self.tabela_mfp479.loc[:,"Local":"Amarelo"].style.set_caption('Consumo de Toner HP LaserJet Pro MFP M479fdw').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()


        self.fromaddr = "sistemas.plural@plural.com.br"
        self.toaddr = "kelvin.rocha@plural.com.br"

        self.msg = MIMEMultipart() 
        self.msg['From'] = self.fromaddr 
        self.msg['To'] = self.toaddr 
        self.msg['Subject'] = "Relatório de Consumo e Suprimentos (Simpress)"
        self.html = """\
        <html>
            <head></head>
            <body>
                <p>Segue Relatórios de todas as impressoras e estoque de suprimentos:</p>
                <p></p>
                {0}
                <p></p>
                <p>Toners MFP 432FDN em estoque : <strong>{1}</strong></p>
                <p></p>
                {2}
                <p></p>
                <p>Toners E52645 em estoque : <strong>{3}</strong></p>
                <p></p>
                {4}
                <p></p>
                <p>Suporte TI - (11) 4152-9518 / 9821</p>
            </body>
        </html>
        """.format(self.relatorio1.to_html(),self.Toner_MFP432.to_string(index=False),self.relatorio2.to_html(),self.Toner_MFPE52645.to_string(index=False),self.relatorio3.to_html())
        self.body = MIMEText(self.html, 'html')
        self.msg.attach(self.body)
        self.filename = "CONTROLE DE IMPRESSORAS.xlsx"
        self.attachment = open(r"C:\Users\kelvin.rocha\Desktop\Area de Trabalho\Kelvin\Python\Automação Python\Breatifulsoap\CONTROLE_DE_IMPRESSORAS.xlsx","rb") 
        self.p = MIMEBase('application', 'octet-stream') 
        self.p.set_payload((self.attachment).read()) 
        encoders.encode_base64(self.p) 

        self.p.add_header('Content-Disposition', "attachment; filename= %s" % self.filename) 
        self.msg.attach(self.p) 
        self.s = smtplib.SMTP('email.plural.com.br') 
        self.s.ehlo()
        self.s.login("Sistemas.plural","asdf321!@#") 
        self.text = self.msg.as_string() 
        self.s.sendmail(self.fromaddr, self.toaddr, self.text) 
        self.s.quit() 
        print('E-mail Enviado com Sucesso!')
        time.sleep(3)

Start = Relatorio_Impressoras()
Start.Inicio()

