import time
from datetime import datetime
import datetime as dt
from selenium import webdriver 
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.message import EmailMessage
from email.mime.base import MIMEBase 
from email import encoders 
import pandas as pd
import smtplib
import openpyxl


print('Iniciando Sistema...')

class Relatorio_Impressoras:
    def Inicio(self):
        self.Agora = dt.datetime.now()
        self.DataAtual = datetime.today()
        self.Hora = '{}'.format(self.DataAtual.hour)
        self.GeradoEm = self.Agora.strftime("%d/%m/%Y %H:%M")
        
        if int(self.Hora) >= 5 and int(self.Hora) <= 7:

            self.planilha = openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
            self.modelo = self.planilha['Modelo']
            self.Hoje = dt.datetime.now()
            self.Ontem = self.Hoje - dt.timedelta(1)
            self.DiaAtual = self.Hoje.strftime("%d-%m-%y")
            self.OntemdeOntem = self.Ontem - dt.timedelta(1)
            print('Data Atual: ',self.DiaAtual)
            self.DiaAnterior = self.Ontem.strftime("%d-%m-%y")  
            print('Data Anterior: ',self.DiaAnterior)
            print('Ontem de Ontem: ',self.OntemdeOntem)
            self.lista = self.planilha.sheetnames
        
            if self.DiaAtual in self.lista:
                print("Sheet Existe!...")
                print('Iniciando Relatorio Primeiro Turno...')
                self.MFP_432_PrimeiroTurno()
                self.Salvar_Dados_MFP_432_PrimeiroTurno()
                self.MFP_E52645_PrimeiroTurno()
                self.Salvar_Dados_MFP_E52645_PrimeiroTurno()
                self.MFP_M479fdw_PrimeiroTurno()
                self.Salvar_Dados_MFP_M479fdw_PrimeiroTurno()
                self.Enviar_Email_PrimeiroTurno()
            else:
                print("Criando a Sheet :",self.DiaAtual)
                self.novaPlanilha = self.planilha.copy_worksheet(self.modelo).title = self.DiaAtual
                self.planilha.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
                print('Iniciando Relatorio Primeiro Turno...')
                self.MFP_432_PrimeiroTurno()
                self.Salvar_Dados_MFP_432_PrimeiroTurno()
                self.MFP_E52645_PrimeiroTurno()
                self.Salvar_Dados_MFP_E52645_PrimeiroTurno()
                self.MFP_M479fdw_PrimeiroTurno()
                self.Salvar_Dados_MFP_M479fdw_PrimeiroTurno()
                self.Enviar_Email_PrimeiroTurno()
        

        elif int(self.Hora) >= 13 and int(self.Hora) <= 15:
            self.planilha = openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
            self.modelo = self.planilha['Modelo']
            self.Hoje = dt.datetime.now()
            self.Ontem = self.Hoje - dt.timedelta(1)
            self.DiaAtual = self.Hoje.strftime("%d-%m-%y")
            print('Data Atual: ',self.DiaAtual)
            self.DiaAnterior = self.Ontem.strftime("%d-%m-%y")  
            print('Data Anterior: ',self.DiaAnterior)
            print("Verificando se existe a Sheet:",self.DiaAtual)
            self.lista = self.planilha.sheetnames

            if self.DiaAtual in self.lista:
                print("Sheet Existe!...")
                print('Iniciando Relatorio Segundo Turno...')
                self.MFP_432_SegundoTurno()
                self.Salvar_Dados_MFP_432_SegundoTurno()
                self.MFP_E52645_SegundoTurno()
                self.Salvar_Dados_MFP_E52645_SegundoTurno()
                self.MFP_M479fdw_SegundoTurno()
                self.Salvar_Dados_MFP_M479fdw_SegundoTurno()
                self.Enviar_Email_SegundoTurno()
            else:
                print("Sheet nao encontrada...")
                print("Criando a Sheet :",self.DiaAtual)
                self.novaPlanilha = self.planilha.copy_worksheet(self.modelo).title = self.DiaAtual
                self.planilha.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
                print('Iniciando Relatorio Segundo Turno...')
                self.MFP_432_SegundoTurno()
                self.Salvar_Dados_MFP_432_SegundoTurno()
                self.MFP_E52645_SegundoTurno()
                self.Salvar_Dados_MFP_E52645_SegundoTurno()
                self.MFP_M479fdw_SegundoTurno()
                self.Salvar_Dados_MFP_M479fdw_SegundoTurno()
                self.Enviar_Email_SegundoTurno()

            
        elif int(self.Hora) >= 21 and int(self.Hora) <= 23:

            self.planilha = openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
            self.modelo = self.planilha['Modelo']
            self.Hoje = dt.datetime.now()
            self.Ontem = self.Hoje - dt.timedelta(1)
            self.DiaAtual = self.Hoje.strftime("%d-%m-%y")
            print('Data Atual: ',self.DiaAtual)
            self.DiaAnterior = self.Ontem.strftime("%d-%m-%y")  
            print('Data Anterior: ',self.DiaAnterior)
            print("Verificando se existe a Sheet:",self.DiaAtual)
            self.lista = self.planilha.sheetnames

            if self.DiaAtual in self.lista:
                print("Sheet Existe!...")
                print('Iniciando Relatorio Terceiro Turno...')
                self.MFP_432_TerceiroTurno()
                self.Salvar_Dados_MFP_432_TerceiroTurno()
                self.MFP_E52645_TerceiroTurno()
                self.Salvar_Dados_MFP_E52645_TerceiroTurno()
                self.MFP_M479fdw_TerceiroTurno()
                self.Salvar_Dados_MFP_M479fdw_TerceiroTurno()
                self.Enviar_Email_TerceiroTurno()
            else:
                print("Sheet nao encontrada...")
                print("Criando a Sheet :",self.DiaAtual)
                self.novaPlanilha = self.planilha.copy_worksheet(self.modelo).title = self.DiaAtual
                self.planilha.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
                print('Iniciando Relatorio Terceiro Turno...')
                self.MFP_432_TerceiroTurno()
                self.Salvar_Dados_MFP_432_TerceiroTurno()
                self.MFP_E52645_TerceiroTurno()
                self.Salvar_Dados_MFP_E52645_TerceiroTurno()
                self.MFP_M479fdw_TerceiroTurno()
                self.Salvar_Dados_MFP_M479fdw_TerceiroTurno()
                self.Enviar_Email_TerceiroTurno()
            
        else:
            print('Iniciando Relatorio Fora de Turno...')            
            #self.MFP_432()
            #self.Salvar_Dados_MFP_432()
            #self.MFP_E52645()
            #self.Salvar_Dados_MFP_E52645()
            self.MFP_M479fdw()
            self.Salvar_Dados_MFP_M479fdw()
            self.Enviar_Email()
            
        #PRIMEIRO TURNO
    def MFP_432_PrimeiroTurno (self): #OK
        self.offline = 0
        print('Modelo MFP 432...')
        self.lista_dos_toner = []
        self.lista_da_imagem = []
        self.nome_impressoras = ['Jurídico', 'Segurança Trabalho',
                                'Ambulatório Dra Renata','TI',
                                'RH','Pré Impressão',
                                'Ambulatório','Comercial',
                                'Papel e Tinta','Expedição',
                                'Manutenção','Portaria',
                                'Produção','Impressão Digital',
                                'Recepção']
        self.modelo_mfp_432 = [
        'https://10.10.4.150/sws/index.html', #Juridico
        'https://10.10.4.173/sws/index.html', #Segurança do Trabalho
        'https://10.10.4.177/sws/index.html', #Ambulatorio Dra Renata
        'https://10.10.4.174/sws/index.html', #TI
        'https://10.10.4.157/sws/index.html', #RH
        'https://10.10.4.151/sws/index.html', #Pré-Impressão
        'https://10.10.4.155/sws/index.html', #Ambulatório
        'https://10.10.4.162/sws/index.html', #Comercial
        'https://10.10.4.164/sws/index.html', #Papel e Tinta
        'https://10.10.4.163/sws/index.html', #Expedição
        'https://10.10.4.158/sws/index.html', #Manutenção
        'https://10.10.4.160/sws/index.html', #Portaria
        'https://10.10.4.156/sws/index.html', #Produção
        'https://10.10.4.169/sws/index.html', #Impressão Digital
        'https://10.10.4.182/sws/index.html', #Recepção
        ]
        print('Abrindo Navegador')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        
        
        time.sleep(3)
        for self.i in range(15):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_mfp_432[self.i])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()                
                print("Procurando Dados...")
                self.toner = 1
                while(self.toner == 1):
                    try:
                        self.toner = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="ext-gen300"]/div/table/tbody/tr/td[2]/div/div/div[2]')))
                        self.lista_dos_toner.append(self.toner.text)
                        time.sleep(2)
                        print('Armazenando dados do Toner...')
                        self.imagem = self.navegador.find_element(By.XPATH,'//*[@id="ext-gen344"]/div/table/tbody/tr/td[2]/div/div/div[2]')
                        self.lista_da_imagem.append(self.imagem.text)
                        time.sleep(2)
                        print('Armazenando dados da Imagem...')    
                        print(self.lista_dos_toner, self.lista_da_imagem)  
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.toner = 1
                
                
            except:
                print("Error: Impressora Offline!")
                self.planilha = openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
                self.lista = self.planilha.sheetnames

                if self.DiaAnterior in self.lista:
                    self.offline += 1
                    self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAnterior)

                    self.TT_432 = self.planilha.iloc[2:17,14:18].rename(columns= {'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo',
                                                                                'Unnamed: 16':f'% de Toner','Unnamed: 17':'Uni.Imagem'})

                    self.Toner = self.TT_432.loc[self.TT_432["Local"] == self.nome_impressoras[self.i],f"% de Toner"]
                    self.Imagem = self.TT_432.loc[self.TT_432["Local"] == self.nome_impressoras[self.i],"Uni.Imagem"]

                    self.lista_dos_toner.append(self.Toner.to_string(index=False))
                    self.lista_da_imagem.append(self.Imagem.to_string(index=False)) 
                    print(self.lista_dos_toner, self.lista_da_imagem) 

                else:
                    self.offline += 1
                    self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.OntemdeOntem)

                    self.TT_432 = self.planilha.iloc[2:17,14:18].rename(columns= {'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo',
                                                                                'Unnamed: 16':f'% de Toner','Unnamed: 17':'Uni.Imagem'})

                    self.Toner = self.TT_432.loc[self.TT_432["Local"] == self.nome_impressoras[self.i],f"% de Toner"]
                    self.Imagem = self.TT_432.loc[self.TT_432["Local"] == self.nome_impressoras[self.i],"Uni.Imagem"]

                    self.lista_dos_toner.append(self.Toner.to_string(index=False))
                    self.lista_da_imagem.append(self.Imagem.to_string(index=False)) 
                    print(self.lista_dos_toner, self.lista_da_imagem) 
        print("Impressoras MFP 432 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_432_PrimeiroTurno (self): #OK
        print('Carregando Planilha...')
        self.index = 4
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.lista_dos_toner, self.lista_da_imagem):
                self.planilha.cell(column=3, row=self.index, value=self.nome)
                self.planilha.cell(column=4, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")         
        print("Planilha Salva com Sucesso")    
    def MFP_E52645_PrimeiroTurno(self): #OK
        print('Modelo MFP E52645...')
        self.cartucho_preto = []
        self.nome_impressoras = ['ALMOXARIFADO','QUALIDADE','ATENDIMENTO','FINANCEIRO','COMPRAS']
        self.kit_aliment_documentos = []
        self.modelo_e52645 = [
        'https://10.10.4.152/hp/device/DeviceStatus/Index', #Almoxarifado
        'https://10.10.4.153/', #Qualidade
        'https://10.10.4.154/', #Atendimento PB
        'https://10.10.4.171/', #Financeiro
        'https://10.10.4.159/', #Compras
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.l in range(5):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_e52645[self.l])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.cartucho = 1
                while(self.cartucho == 1):
                    try:
                        self.cartucho = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="SupplyPLR0"]')))
                        print('Armazenando dados do Cartucho...')
                        self.cartuchop = self.cartucho.text.split('*')
                        self.cartucho_preto.append(self.cartuchop[0])
                        self.kit = self.navegador.find_element(By.XPATH,'//*[@id="SupplyPLR1"]')
                        print('Armazenado dados Do kit...')
                        self.kitA = self.kit.text.split('*')
                        self.kit_aliment_documentos.append(self.kitA[0])
                        time.sleep(3)
                        print(self.cartucho_preto, self.kit_aliment_documentos)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.cartucho = 1
            except:
                print("Error: Impressora Offline!")
                self.planilha = openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
                self.lista = self.planilha.sheetnames
                if self.DiaAnterior in self.lista:
                    self.offline += 1
                    self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAnterior)

                    self.TT_E52645 = self.planilha.iloc[19:24,14:18].rename(columns= {'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo',
                                                                                'Unnamed: 16':'Toner Preto','Unnamed: 17':'Kit Aliment. Doc.'})

                    self.Toner = self.TT_E52645.loc[self.TT_E52645["Local"] == self.nome_impressoras[self.i],"Toner Preto"]                                                                 
                    self.Kit = self.TT_E52645.loc[self.TT_E52645["Local"] == self.nome_impressoras[self.i],"Kit Aliment. Doc."]

                    self.lista_dos_toner.append(self.Toner.to_string(index=False))
                    self.lista_da_imagem.append(self.Kit.to_string(index=False))
                    print(self.lista_dos_toner, self.lista_da_imagem)
                else:
                    self.offline += 1
                    self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.OntemdeOntem)

                    self.TT_E52645 = self.planilha.iloc[19:24,14:18].rename(columns= {'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo',
                                                                                'Unnamed: 16':'Toner Preto','Unnamed: 17':'Kit Aliment. Doc.'})

                    self.Toner = self.TT_E52645.loc[self.TT_E52645["Local"] == self.nome_impressoras[self.i],"Toner Preto"]                                                                 
                    self.Kit = self.TT_E52645.loc[self.TT_E52645["Local"] == self.nome_impressoras[self.i],"Kit Aliment. Doc."]

                    self.lista_dos_toner.append(self.Toner.to_string(index=False))
                    self.lista_da_imagem.append(self.Kit.to_string(index=False))
                    print(self.lista_dos_toner, self.lista_da_imagem)
        print("Impressoras E52645 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_E52645_PrimeiroTurno(self): #OK
        print('Carregando Planilha...')
        self.index = 21
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.cartucho_preto, self.kit_aliment_documentos):
                self.planilha.cell(column=3, row=self.index, value=self.nome)
                self.planilha.cell(column=4, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_M479fdw_PrimeiroTurno(self): #OK
        print('Modelo MFP M479fdw...')
        self.cartucho_preto = []
        self.cartucho_cyan = []
        self.cartucho_magenta = []
        self.cartucho_yellow = []
        self.nome_impressoras = ['Dir. Carlos Jacomine', 'Anatilia',
                                'Atendimento Color','Rogerio Cordeiro',
                                'Welinton Martins','Adriana Gasparine']
        self.modelo_M479fdw = [
        'https://10.10.4.178/#hId-pgConsumables',
        'https://10.10.4.167/#hId-pgConsumables',
        'https://10.10.4.168/#hId-pgConsumables',
        'https://10.10.4.175/#hId-pgConsumables',
        'https://10.10.4.172/#hId-pgConsumables',
        'https://10.10.4.170/#hId-pgConsumables'
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.j in range(6):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_M479fdw[self.j])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.preto = 1
                while(self.preto == 1):
                    try:
                        self.preto = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[2]')))
                        self.cyan = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[3]')
                        self.magenta = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[4]')
                        self.yellow = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[5]')
                        print('Armazenando dados dos Cartuchos...')
                        self.cartucho_preto.append(self.preto.text)
                        self.cartucho_cyan.append(self.cyan.text)
                        self.cartucho_magenta.append(self.magenta.text)
                        self.cartucho_yellow.append(self.yellow.text)
                        time.sleep(3)
                        print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.preto = 1
            except:
                print("Error: Impressora Offline!")
                self.planilha = openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
                self.lista = self.planilha.sheetnames
                if self.DiaAnterior in self.lista:
                    self.offline += 1
                    self.planilha = pd.read_excel('Relatorio_Diario.xlsx',sheet_name= self.DiaAnterior)

                    self.TT_479 = self.planilha.iloc[26:32,14:20].rename(columns= {'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo',
                                                                            'Unnamed: 16':'Preto','Unnamed: 17':'Ciano','Unnamed: 18':'Magenta','Unnamed: 19':'Amarelo'})

                    self.Preto = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Preto"]                                                                 
                    self.Ciano = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Ciano"]
                    self.Magenta = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Magenta"]                                                                 
                    self.Amarelo = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Amarelo"]                
    
                    self.cartucho_preto.append(self.Preto.to_string(index=False))
                    self.cartucho_cyan.append(self.Ciano.to_string(index=False))
                    self.cartucho_magenta.append(self.Magenta.to_string(index=False))
                    self.cartucho_yellow.append(self.Amarelo.to_string(index=False))
                    print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
                else:
                    self.offline += 1
                    self.planilha = pd.read_excel('Relatorio_Diario.xlsx',sheet_name= self.OntemdeOntem)

                    self.TT_479 = self.planilha.iloc[26:32,14:20].rename(columns= {'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo',
                                                                            'Unnamed: 16':'Preto','Unnamed: 17':'Ciano','Unnamed: 18':'Magenta','Unnamed: 19':'Amarelo'})

                    self.Preto = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Preto"]                                                                 
                    self.Ciano = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Ciano"]
                    self.Magenta = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Magenta"]                                                                 
                    self.Amarelo = self.TT_479.loc[self.TT_479["Local"] == self.nome_impressoras[self.i],"Amarelo"]                
    
                    self.cartucho_preto.append(self.Preto.to_string(index=False))
                    self.cartucho_cyan.append(self.Ciano.to_string(index=False))
                    self.cartucho_magenta.append(self.Magenta.to_string(index=False))
                    self.cartucho_yellow.append(self.Amarelo.to_string(index=False))
                    print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
        print("Impressoras MFP M479fdw Offline: ",self.offline)
        self.offline = 0                
        print('Fechando Navegador...')
        self.navegador.quit()      
    def Salvar_Dados_MFP_M479fdw_PrimeiroTurno(self): #OK
        print('Carregando Planilha...')
        self.index = 28
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.colorblack, self.colorcyan, self.colormagenta, self.coloryellow in zip(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow):
                self.planilha.cell(column=3, row=self.index, value=self.colorblack)
                self.planilha.cell(column=4, row=self.index, value=self.colorcyan)
                self.planilha.cell(column=5, row=self.index, value=self.colormagenta)
                self.planilha.cell(column=6, row=self.index, value=self.coloryellow)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def Enviar_Email_PrimeiroTurno(self): #OK
        self.tipos_colunas = {'Modelo' : str,'Estoque': str}
        self.lista_colunas = ['Modelo','Cor','Estoque']
        self.tabela_simpress = pd.read_excel(r'\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx', sheet_name='Python')
        pd.set_option('display.precision',0)
        self.Toner_MFP432 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "432FDN","Estoque"]
        self.Toner_MFPE52645 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "E52645","Estoque"]
        self.Toner_MFP7E77830 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "HP 7E77830","Estoque"]
        self.M479_Black = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Black","Estoque"]
        self.M479_Cyan = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Cyan","Estoque"]
        self.M479_Pink = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Pink","Estoque"]
        self.M479_Yellow = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Yellow","Estoque"]

        print('Enviando E-mail...')
        time.sleep(3)

        self.planilhaAberta = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

        self.relatorio1 = self.planilhaAberta.iloc[2:17,:4].rename(columns={'PRIMEIRO TURNO':'Local','Unnamed: 1':'Modelo','Unnamed: 2':f'% de Toner','Unnamed: 3':'Uni.Imagem'}).style.set_caption('Consumo de Toner HP Laser MFP 432').set_table_styles([
             { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
              'padding-right': '20px'}).hide_index()
        self.relatorio2 = self.planilhaAberta.iloc[19:24,:4].rename(columns={'PRIMEIRO TURNO':'Local','Unnamed: 1':'Modelo','Unnamed: 2':'Toner Preto','Unnamed: 3':'Kit Aliment. Doc.'}).style.set_caption('Consumo de Toner HP LaserJet MFP E52645').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.relatorio3 = self.planilhaAberta.iloc[26:32,:6].rename(columns={'PRIMEIRO TURNO':'Local','Unnamed: 1':'Modelo','Unnamed: 2':'Preto','Unnamed: 3':'Ciano','Unnamed: 4':'Magenta','Unnamed: 5':'Amarelo'}).style.set_caption('Consumo de Toner HP LaserJet Pro MFP M479fdw').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        
        self.fromaddr = "sistemas.plural@plural.com.br"
        self.toaddr = "suporte@plural.com.br"
        self.msg = MIMEMultipart() 
        self.msg['From'] = self.fromaddr 
        self.msg['To'] = self.toaddr 
        self.msg['Subject'] = "Primeiro Turno - Relatório (Simpress)"
        self.html = """\
        <html>
            <head></head>
            <body>
                <h3 style="margin-left:140px">PRIMEIRO TURNO</h3>
                <p>Relatório gerado em: <strong>{0}</strong></p>
                <p>Segue Relatórios de todas as impressoras e estoque de suprimentos do 1º Turno:</p>
                <p></p>
                {1}
                <p></p>
                <p>Toners MFP 432FDN em estoque : <strong>{2}</strong></p>
                <p></p>
                {3}
                <p></p>
                <p>Toners E52645 em estoque : <strong>{4}</strong></p>
                <p></p>
                {5}
                <p></p>
                <p>
                Toner M479 <strong>Preto</strong> em estoque : <strong>{6}</strong></br>
                Toner M479 <strong>Ciano</strong> em estoque : <strong>{7}</strong></br>
                Toner M479 <strong>Magenta</strong> em estoque : <strong>{8}</strong></br>
                Toner M479 <strong>Amarelo</strong> em estoque : <strong>{9}</strong></br>
                </p>
                <p></p>
                <p>Abra a planilha dos Relatórios Diarios <a href="\\\srvsao028\Automação Python\Relatorio_Diario.xlsx"
                                                            target="_blank">Clicando aqui</a>
                </p>
                <p>Abra a planilha de Controle de Estoque Simpress <a href="\\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx""
                                                                    target="_blank">Clicando aqui</a>
                </p>
            <p>Suporte TI - (11) 4152-9518 / 9821</p>
            </body>
        </html>
        """.format(self.GeradoEm,self.relatorio1.to_html(),self.Toner_MFP432.to_string(index=False),self.relatorio2.to_html(),self.Toner_MFPE52645.to_string(index=False),self.relatorio3.to_html(),self.M479_Black.to_string(index=False),self.M479_Cyan.to_string(index=False),self.M479_Pink.to_string(index=False),self.M479_Yellow.to_string(index=False))
        self.body = MIMEText(self.html, 'html')
        self.msg.attach(self.body)
        self.filename = "Relatorio_Diario.xlsx"
        self.attachment = open(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx","rb") 
        self.p = MIMEBase('application', 'octet-stream') 
        self.p.set_payload((self.attachment).read()) 
        encoders.encode_base64(self.p) 
        self.p.add_header('Content-Disposition', "attachment; filename= %s" % self.filename) 
        self.msg.attach(self.p) 
        self.s = smtplib.SMTP('email.plural.com.br') 
        self.s.ehlo()
        #self.s.login("Sistemas.plural","asdf321!@#") 
        self.text = self.msg.as_string() 
        self.s.sendmail(self.fromaddr, self.toaddr, self.text) 
        self.s.quit() 
        print('E-mail Enviado com Sucesso!')
        time.sleep(3)

        #SEGUNDO TURNO
    def MFP_432_SegundoTurno (self): #OK
        self.offline = 0
        print('Modelo MFP 432...')
        self.lista_dos_toner = []
        self.lista_da_imagem = []
        self.nome_impressoras = ['Jurídico', 'Segurança Trabalho',
                                'Ambulatório Dra Renata','TI',
                                'RH','Pré Impressão',
                                'Ambulatório','Comercial',
                                'Papel e Tinta','Expedição',
                                'Manutenção','Portaria',
                                'Produção','Impressão Digital',
                                'Recepção']
        self.modelo_mfp_432 = [
        'https://10.10.4.150/sws/index.html', #Juridico
        'https://10.10.4.173/sws/index.html', #Segurança do Trabalho
        'https://10.10.4.177/sws/index.html', #Ambulatorio Dra Renata
        'https://10.10.4.174/sws/index.html', #TI
        'https://10.10.4.157/sws/index.html', #RH
        'https://10.10.4.151/sws/index.html', #Pré-Impressão
        'https://10.10.4.155/sws/index.html', #Ambulatório
        'https://10.10.4.162/sws/index.html', #Comercial
        'https://10.10.4.164/sws/index.html', #Papel e Tinta
        'https://10.10.4.163/sws/index.html', #Expedição
        'https://10.10.4.158/sws/index.html', #Manutenção
        'https://10.10.4.160/sws/index.html', #Portaria
        'https://10.10.4.156/sws/index.html', #Produção
        'https://10.10.4.169/sws/index.html', #Impressão Digital
        'https://10.10.4.182/sws/index.html', #Recepção
        ]
        print('Abrindo Navegador')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.i in range(15):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_mfp_432[self.i])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.toner = 1
                while(self.toner == 1):
                    try:
                        self.toner = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="ext-gen300"]/div/table/tbody/tr/td[2]/div/div/div[2]')))
                        self.lista_dos_toner.append(self.toner.text)
                        time.sleep(2)
                        print('Armazenando dados do Toner...')
                        self.imagem = self.navegador.find_element(By.XPATH,'//*[@id="ext-gen344"]/div/table/tbody/tr/td[2]/div/div/div[2]')
                        self.lista_da_imagem.append(self.imagem.text)
                        time.sleep(2)
                        print('Armazenando dados da Imagem...')    
                        print(self.lista_dos_toner, self.lista_da_imagem)  
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.toner = 1        
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

                self.PT_432 = self.planilha.iloc[2:17,:4].rename(columns= {'PRIMEIRO TURNO':'Local','Unnamed: 1':'Modelo',
                                                                'Unnamed: 2':f'% de Toner','Unnamed: 3':'Uni.Imagem'})

                self.Toner = self.PT_432.loc[self.PT_432["Local"] == self.nome_impressoras[self.i],f"% de Toner"]
                self.Imagem = self.PT_432.loc[self.PT_432["Local"] == self.nome_impressoras[self.i],"Uni.Imagem"]

                self.lista_dos_toner.append(self.Toner.to_string(index=False))
                self.lista_da_imagem.append(self.Imagem.to_string(index=False)) 
                print(self.lista_dos_toner, self.lista_da_imagem) 
        print("Impressoras MFP 432 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_432_SegundoTurno (self): #OK
        print('Carregando Planilha...')
        self.index = 4
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.lista_dos_toner, self.lista_da_imagem):
                self.planilha.cell(column=10, row=self.index, value=self.nome)
                self.planilha.cell(column=11, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_E52645_SegundoTurno(self): #OK
        print('Modelo MFP E52645...')
        self.cartucho_preto = []
        self.nome_impressoras = ['ALMOXARIFADO','QUALIDADE','ATENDIMENTO','FINANCEIRO','COMPRAS']
        self.kit_aliment_documentos = []
        self.modelo_e52645 = [
        'https://10.10.4.152/hp/device/DeviceStatus/Index', #Almoxarifado
        'https://10.10.4.153/', #Qualidade
        'https://10.10.4.154/', #Atendimento PB
        'https://10.10.4.171/', #Financeiro
        'https://10.10.4.159/', #Compras
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.l in range(5):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_e52645[self.l])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.cartucho = 1
                while(self.cartucho == 1):
                    try:
                        self.cartucho = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="SupplyPLR0"]')))
                        print('Armazenando dados do Cartucho...')
                        self.cartuchop = self.cartucho.text.split('*')
                        self.cartucho_preto.append(self.cartuchop[0])
                        self.kit = self.navegador.find_element(By.XPATH,'//*[@id="SupplyPLR1"]')
                        print('Armazenado dados Do kit...')
                        self.kitA = self.kit.text.split('*')
                        self.kit_aliment_documentos.append(self.kitA[0])
                        time.sleep(3)
                        print(self.cartucho_preto, self.kit_aliment_documentos)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.cartucho = 1
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

                self.PT_E52645 = self.planilha.iloc[19:24,:4].rename(columns= {'PRIMEIRO TURNO':'Local','Unnamed: 1':'Modelo',
                                                                    'Unnamed: 2':'Toner Preto','Unnamed: 3':'Kit Aliment. Doc.'})

                self.Toner = self.PT_E52645.loc[self.PT_E52645["Local"] == self.nome_impressoras[self.i],"Toner Preto"]                                                                 
                self.Kit = self.PT_E52645.loc[self.PT_E52645["Local"] == self.nome_impressoras[self.i],"Kit Aliment. Doc."]

                self.cartucho_preto.append(self.Toner.to_string(index=False))
                self.kit_aliment_documentos.append(self.Kit.to_string(index=False))
                print(self.cartucho_preto, self.kit_aliment_documentos) 
        print("Impressoras E52645 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_E52645_SegundoTurno(self): #OK
        print('Carregando Planilha...')
        self.index = 21
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.cartucho_preto, self.kit_aliment_documentos):
                self.planilha.cell(column=10, row=self.index, value=self.nome)
                self.planilha.cell(column=11, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_M479fdw_SegundoTurno(self): #OK
        print('Modelo MFP M479fdw...')
        self.cartucho_preto = []
        self.cartucho_cyan = []
        self.cartucho_magenta = []
        self.cartucho_yellow = []
        self.nome_impressoras = ['Dir. Carlos Jacomine', 'Anatilia',
                                'Atendimento Color','Rogerio Cordeiro',
                                'Welinton Martins','Adriana Gasparine']
        self.modelo_M479fdw = [
        'https://10.10.4.178/#hId-pgConsumables',
        'https://10.10.4.167/#hId-pgConsumables',
        'https://10.10.4.168/#hId-pgConsumables',
        'https://10.10.4.175/#hId-pgConsumables',
        'https://10.10.4.172/#hId-pgConsumables',
        'https://10.10.4.170/#hId-pgConsumables'
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.j in range(6):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_M479fdw[self.j])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.preto = 1
                while(self.preto == 1):
                    try:
                        self.preto = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[2]')))
                        self.cyan = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[3]')
                        self.magenta = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[4]')
                        self.yellow = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[5]')
                        print('Armazenando dados dos Cartuchos...')
                        self.cartucho_preto.append(self.preto.text)
                        self.cartucho_cyan.append(self.cyan.text)
                        self.cartucho_magenta.append(self.magenta.text)
                        self.cartucho_yellow.append(self.yellow.text)
                        time.sleep(3)
                        print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.preto = 1
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

                self.PT_479 = self.planilha.iloc[26:32,:6].rename(columns= {'PRIMEIRO TURNO':'Local','Unnamed: 1':'Modelo',
                                                                    'Unnamed: 2':'Preto','Unnamed: 3':'Ciano','Unnamed: 4':'Magenta','Unnamed: 5':'Amarelo'})

                self.Preto = self.PT_479.loc[self.PT_479["Local"] == self.nome_impressoras[self.i],"Preto"]                                                                 
                self.Ciano = self.PT_479.loc[self.PT_479["Local"] == self.nome_impressoras[self.i],"Ciano"]
                self.Magenta = self.PT_479.loc[self.PT_479["Local"] == self.nome_impressoras[self.i],"Magenta"]                                                                 
                self.Amarelo = self.PT_479.loc[self.PT_479["Local"] == self.nome_impressoras[self.i],"Amarelo"]
 
                self.cartucho_preto.append(self.Preto.to_string(index=False))
                self.cartucho_cyan.append(self.Ciano.to_string(index=False))
                self.cartucho_magenta.append(self.Magenta.to_string(index=False))
                self.cartucho_yellow.append(self.Amarelo.to_string(index=False))
                print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow) 
        print("Impressoras MFP M479fdw Offline: ",self.offline)
        self.offline = 0                
        print('Fechando Navegador...')
        self.navegador.quit()   
    def Salvar_Dados_MFP_M479fdw_SegundoTurno(self): #OK
        print('Carregando Planilha...')
        self.index = 28
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.colorblack, self.colorcyan, self.colormagenta, self.coloryellow in zip(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow):
                self.planilha.cell(column=10, row=self.index, value=self.colorblack)
                self.planilha.cell(column=11, row=self.index, value=self.colorcyan)
                self.planilha.cell(column=12, row=self.index, value=self.colormagenta)
                self.planilha.cell(column=13, row=self.index, value=self.coloryellow)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def Enviar_Email_SegundoTurno(self): #OK
        self.tipos_colunas = {'Modelo' : str,'Estoque': str}
        self.lista_colunas = ['Modelo','Cor','Estoque']
        self.tabela_simpress = pd.read_excel(r'\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx', sheet_name='Python')
        pd.set_option('display.precision',0)
        self.Toner_MFP432 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "432FDN","Estoque"]
        self.Toner_MFPE52645 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "E52645","Estoque"]
        self.Toner_MFP7E77830 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "HP 7E77830","Estoque"]
        self.M479_Black = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Black","Estoque"]
        self.M479_Cyan = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Cyan","Estoque"]
        self.M479_Pink = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Pink","Estoque"]
        self.M479_Yellow = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Yellow","Estoque"]

        print('Enviando E-mail...')
        time.sleep(3)

        self.planilhaAberta = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

        self.relatorio1 = self.planilhaAberta.iloc[2:17,7:11].rename(columns={'SEGUNDO TURNO':'Local','Unnamed: 8':'Modelo','Unnamed: 9':f'% de Toner','Unnamed: 10':'Uni.Imagem'}).style.set_caption('Consumo de Toner HP Laser MFP 432').set_table_styles([
             { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
              'padding-right': '20px'}).hide_index()
        self.relatorio2 = self.planilhaAberta.iloc[19:24,7:11].rename(columns={'SEGUNDO TURNO':'Local','Unnamed: 8':'Modelo','Unnamed: 9':'Toner Preto','Unnamed: 10':'Kit Aliment. Doc.'}).style.set_caption('Consumo de Toner HP LaserJet MFP E52645').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.relatorio3 = self.planilhaAberta.iloc[26:32,7:13].rename(columns={'SEGUNDO TURNO':'Local','Unnamed: 8':'Modelo','Unnamed: 9':'Preto','Unnamed: 10':'Ciano','Unnamed: 11':'Magenta','Unnamed: 12':'Amarelo'}).style.set_caption('Consumo de Toner HP LaserJet Pro MFP M479fdw').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.fromaddr = "sistemas.plural@plural.com.br"
        self.toaddr = "suporte@plural.com.br"
        self.msg = MIMEMultipart() 
        self.msg['From'] = self.fromaddr 
        self.msg['To'] = self.toaddr 
        self.msg['Subject'] = "Segundo Turno - Relatório (Simpress)"
        self.html = """\
        <html>
            <head></head>
            <body>
                <h3 style="margin-left:140px">SEGUNDO TURNO</h3>
                <p>Relatório gerado em: <strong>{0}</strong></p>
                <p>Segue Relatórios de todas as impressoras e estoque de suprimentos do 2º Turno:</p>
                <p></p>
                {1}
                <p></p>
                <p>Toners MFP 432FDN em estoque : <strong>{2}</strong></p>
                <p></p>
                {3}
                <p></p>
                <p>Toners E52645 em estoque : <strong>{4}</strong></p>
                <p></p>
                {5}
                <p></p>
                <p>
                Toner M479 <strong>Preto</strong> em estoque : <strong>{6}</strong></br>
                Toner M479 <strong>Ciano</strong> em estoque : <strong>{7}</strong></br>
                Toner M479 <strong>Magenta</strong> em estoque : <strong>{8}</strong></br>
                Toner M479 <strong>Amarelo</strong> em estoque : <strong>{9}</strong></br>
                </p>
                <p></p>
                <p>Abra a planilha dos Relatórios Diarios <a href="\\\srvsao028\Automação Python\Relatorio_Diario.xlsx"
                                                            target="_blank">Clicando aqui</a>
                </p>
                <p>Abra a planilha de Controle de Estoque Simpress <a href="\\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx""
                                                                    target="_blank">Clicando aqui</a>
                </p>
            <p>Suporte TI - (11) 4152-9518 / 9821</p>
            </body>
        </html>
        """.format(self.GeradoEm,self.relatorio1.to_html(),self.Toner_MFP432.to_string(index=False),self.relatorio2.to_html(),self.Toner_MFPE52645.to_string(index=False),self.relatorio3.to_html(),self.M479_Black.to_string(index=False),self.M479_Cyan.to_string(index=False),self.M479_Pink.to_string(index=False),self.M479_Yellow.to_string(index=False))
        self.body = MIMEText(self.html, 'html')
        self.msg.attach(self.body)
        self.filename = "Relatorio_Diario.xlsx"
        self.attachment = open(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx","rb") 
        self.p = MIMEBase('application', 'octet-stream') 
        self.p.set_payload((self.attachment).read()) 
        encoders.encode_base64(self.p) 
        self.p.add_header('Content-Disposition', "attachment; filename= %s" % self.filename) 
        self.msg.attach(self.p) 
        self.s = smtplib.SMTP('email.plural.com.br') 
        self.s.ehlo()
        #self.s.login("Sistemas.plural","asdf321!@#") 
        self.text = self.msg.as_string() 
        self.s.sendmail(self.fromaddr, self.toaddr, self.text) 
        self.s.quit() 
        print('E-mail Enviado com Sucesso!')
        time.sleep(3)


        #TERCEIRO TURNO
    def MFP_432_TerceiroTurno (self): #OK
        self.offline = 0
        print('Modelo MFP 432...')
        self.lista_dos_toner = []
        self.lista_da_imagem = []
        self.nome_impressoras = ['Jurídico', 'Segurança Trabalho',
                                'Ambulatório Dra Renata','TI',
                                'RH','Pré Impressão',
                                'Ambulatório','Comercial',
                                'Papel e Tinta','Expedição',
                                'Manutenção','Portaria',
                                'Produção','Impressão Digital',
                                'Recepção']
        self.modelo_mfp_432 = [
        'https://10.10.4.150/sws/index.html', #Juridico
        'https://10.10.4.173/sws/index.html', #Segurança do Trabalho
        'https://10.10.4.177/sws/index.html', #Ambulatorio Dra Renata
        'https://10.10.4.174/sws/index.html', #TI
        'https://10.10.4.157/sws/index.html', #RH
        'https://10.10.4.151/sws/index.html', #Pré-Impressão
        'https://10.10.4.155/sws/index.html', #Ambulatório
        'https://10.10.4.162/sws/index.html', #Comercial
        'https://10.10.4.164/sws/index.html', #Papel e Tinta
        'https://10.10.4.163/sws/index.html', #Expedição
        'https://10.10.4.158/sws/index.html', #Manutenção
        'https://10.10.4.160/sws/index.html', #Portaria
        'https://10.10.4.156/sws/index.html', #Produção
        'https://10.10.4.169/sws/index.html', #Impressão Digital
        'https://10.10.4.182/sws/index.html', #Recepção
        ]
        print('Abrindo Navegador')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.i in range(15):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_mfp_432[self.i])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.toner = 1
                while(self.toner == 1):
                    try:
                        self.toner = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="ext-gen300"]/div/table/tbody/tr/td[2]/div/div/div[2]')))
                        self.lista_dos_toner.append(self.toner.text)
                        time.sleep(2)
                        print('Armazenando dados do Toner...')
                        self.imagem = self.navegador.find_element(By.XPATH,'//*[@id="ext-gen344"]/div/table/tbody/tr/td[2]/div/div/div[2]')
                        self.lista_da_imagem.append(self.imagem.text)
                        time.sleep(2)
                        print('Armazenando dados da Imagem...')    
                        print(self.lista_dos_toner, self.lista_da_imagem)  
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.toner = 1          
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

                self.ST_432 = self.planilha.iloc[2:17,7:11].rename(columns= {'SEGUNDO TURNO':'Local','Unnamed: 8':'Modelo',
                                                                            'Unnamed: 9':f'% de Toner','Unnamed: 10':'Uni.Imagem'})

                self.Toner = self.ST_432.loc[self.ST_432["Local"] == self.nome_impressoras[self.i],f"% de Toner"]
                self.Imagem = self.ST_432.loc[self.ST_432["Local"] == self.nome_impressoras[self.i],"Uni.Imagem"]

                self.lista_dos_toner.append(self.Toner.to_string(index=False))
                self.lista_da_imagem.append(self.Imagem.to_string(index=False)) 
                print(self.lista_dos_toner, self.lista_da_imagem) 
        print("Impressoras MFP 432 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_432_TerceiroTurno (self): #OK
        print('Carregando Planilha...')
        self.index = 4
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.lista_dos_toner, self.lista_da_imagem):
                self.planilha.cell(column=17, row=self.index, value=self.nome)
                self.planilha.cell(column=18, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_E52645_TerceiroTurno(self): #OK
        print('Modelo MFP E52645...')
        self.cartucho_preto = []
        self.nome_impressoras = ['ALMOXARIFADO','QUALIDADE','ATENDIMENTO','FINANCEIRO','COMPRAS']
        self.kit_aliment_documentos = []
        self.modelo_e52645 = [
        'https://10.10.4.152/hp/device/DeviceStatus/Index', #Almoxarifado
        'https://10.10.4.153/', #Qualidade
        'https://10.10.4.154/', #Atendimento PB
        'https://10.10.4.171/', #Financeiro
        'https://10.10.4.159/', #Compras
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.l in range(5):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_e52645[self.l])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.cartucho = 1
                while(self.cartucho == 1):
                    try:
                        self.cartucho = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="SupplyPLR0"]')))
                        print('Armazenando dados do Cartucho...')
                        self.cartuchop = self.cartucho.text.split('*')
                        self.cartucho_preto.append(self.cartuchop[0])
                        self.kit = self.navegador.find_element(By.XPATH,'//*[@id="SupplyPLR1"]')
                        print('Armazenado dados Do kit...')
                        self.kitA = self.kit.text.split('*')
                        self.kit_aliment_documentos.append(self.kitA[0])
                        time.sleep(3)
                        print(self.cartucho_preto, self.kit_aliment_documentos)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.cartucho = 1
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

                self.ST_E52645 = self.planilha.iloc[19:24,7:11].rename(columns= {'SEGUNDO TURNO':'Local','Unnamed: 8':'Modelo',
                                                                        'Unnamed: 9':'Toner Preto','Unnamed: 10':'Kit Aliment. Doc.'})

                self.Toner = self.ST_E52645.loc[self.ST_E52645["Local"] == self.nome_impressoras[self.i],"Toner Preto"]                                                                 
                self.Kit = self.ST_E52645.loc[self.ST_E52645["Local"] == self.nome_impressoras[self.i],"Kit Aliment. Doc."]               

                self.cartucho_preto.append(self.Toner.to_string(index=False))
                self.kit_aliment_documentos.append(self.Kit.to_string(index=False))
                print(self.cartucho_preto, self.kit_aliment_documentos) 
        print("Impressoras E52645 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_E52645_TerceiroTurno(self): #OK
        print('Carregando Planilha...')
        self.index = 21
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.cartucho_preto, self.kit_aliment_documentos):
                self.planilha.cell(column=17, row=self.index, value=self.nome)
                self.planilha.cell(column=18, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_M479fdw_TerceiroTurno(self): #OK
        print('Modelo MFP M479fdw...')
        self.cartucho_preto = []
        self.cartucho_cyan = []
        self.cartucho_magenta = []
        self.cartucho_yellow = []
        self.nome_impressoras = ['Dir. Carlos Jacomine', 'Anatilia',
                                'Atendimento Color','Rogerio Cordeiro',
                                'Welinton Martins','Adriana Gasparine']
        self.modelo_M479fdw = [
        'https://10.10.4.178/#hId-pgConsumables',
        'https://10.10.4.167/#hId-pgConsumables',
        'https://10.10.4.168/#hId-pgConsumables',
        'https://10.10.4.175/#hId-pgConsumables',
        'https://10.10.4.172/#hId-pgConsumables',
        'https://10.10.4.170/#hId-pgConsumables'
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        time.sleep(3)
        for self.j in range(6):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_M479fdw[self.j])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print("Procurando Dados...")
                self.preto = 1
                while(self.preto == 1):
                    try:
                        self.preto = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[2]')))
                        self.cyan = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[3]')
                        self.magenta = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[4]')
                        self.yellow = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[5]')
                        print('Armazenando dados dos Cartuchos...')
                        self.cartucho_preto.append(self.preto.text)
                        self.cartucho_cyan.append(self.cyan.text)
                        self.cartucho_magenta.append(self.magenta.text)
                        self.cartucho_yellow.append(self.yellow.text)
                        time.sleep(3)
                        print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.preto = 1
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

                self.ST_479 = self.planilha.iloc[26:32,7:13].rename(columns= {'SEGUNDO TURNO':'Local','Unnamed: 8':'Modelo',
                                                                            'Unnamed: 9':'Preto','Unnamed: 10':'Ciano','Unnamed: 11':'Magenta','Unnamed: 12':'Amarelo'})

                self.Preto = self.ST_479.loc[self.ST_479["Local"] == self.nome_impressoras[self.i],"Preto"]                                                                 
                self.Ciano = self.ST_479.loc[self.ST_479["Local"] == self.nome_impressoras[self.i],"Ciano"]
                self.Magenta = self.ST_479.loc[self.ST_479["Local"] == self.nome_impressoras[self.i],"Magenta"]                                                                 
                self.Amarelo = self.ST_479.loc[self.ST_479["Local"] == self.nome_impressoras[self.i],"Amarelo"]
 
                self.cartucho_preto.append(self.Preto.to_string(index=False))
                self.cartucho_cyan.append(self.Ciano.to_string(index=False))
                self.cartucho_magenta.append(self.Magenta.to_string(index=False))
                self.cartucho_yellow.append(self.Amarelo.to_string(index=False))
                print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow) 
        print("Impressoras MFP M479fdw Offline: ",self.offline)
        self.offline = 0                
        print('Fechando Navegador...')
        self.navegador.quit()    
    def Salvar_Dados_MFP_M479fdw_TerceiroTurno(self): #OK
        print('Carregando Planilha...')
        self.index = 28
        self.book= openpyxl.load_workbook(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")
        self.planilha = self.book[self.DiaAtual]
        for self.i in range(1):
            for self.colorblack, self.colorcyan, self.colormagenta, self.coloryellow in zip(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow):
                self.planilha.cell(column=17, row=self.index, value=self.colorblack)
                self.planilha.cell(column=18, row=self.index, value=self.colorcyan)
                self.planilha.cell(column=19, row=self.index, value=self.colormagenta)
                self.planilha.cell(column=20, row=self.index, value=self.coloryellow)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx")        
        print("Planilha Salva com Sucesso")
    def Enviar_Email_TerceiroTurno(self): #OK
        self.tipos_colunas = {'Modelo' : str,'Estoque': str}
        self.lista_colunas = ['Modelo','Cor','Estoque']
        self.tabela_simpress = pd.read_excel(r'\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx', sheet_name='Python')
        pd.set_option('display.precision',0)
        self.Toner_MFP432 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "432FDN","Estoque"]
        self.Toner_MFPE52645 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "E52645","Estoque"]
        self.Toner_MFP7E77830 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "HP 7E77830","Estoque"]
        self.M479_Black = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Black","Estoque"]
        self.M479_Cyan = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Cyan","Estoque"]
        self.M479_Pink = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Pink","Estoque"]
        self.M479_Yellow = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Yellow","Estoque"]

        print('Enviando E-mail...')
        time.sleep(3)

        self.planilhaAberta = pd.read_excel(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx",sheet_name= self.DiaAtual)

        self.relatorio1 = self.planilhaAberta.iloc[2:17,14:18].rename(columns={'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo','Unnamed: 16':f'% de Toner','Unnamed: 17':'Uni.Imagem'}).style.set_caption('Consumo de Toner HP Laser MFP 432').set_table_styles([
             { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
              'padding-right': '20px'}).hide_index()
        self.relatorio2 = self.planilhaAberta.iloc[19:24,14:18].rename(columns={'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo','Unnamed: 16':'Toner Preto','Unnamed: 17':'Kit Aliment. Doc.'}).style.set_caption('Consumo de Toner HP LaserJet MFP E52645').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.relatorio3 = self.planilhaAberta.iloc[26:32,14:20].rename(columns={'TERCEIRO TURNO':'Local','Unnamed: 15':'Modelo','Unnamed: 16':'Preto','Unnamed: 17':'Ciano','Unnamed: 18':'Magenta','Unnamed: 19':'Amarelo'}).style.set_caption('Consumo de Toner HP LaserJet Pro MFP M479fdw').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.fromaddr = "sistemas.plural@plural.com.br"
        self.toaddr = "suporte@plural.com.br"
        self.msg = MIMEMultipart() 
        self.msg['From'] = self.fromaddr 
        self.msg['To'] = self.toaddr 
        self.msg['Subject'] = "Terceiro Turno - Relatório (Simpress)"
        self.html = """\
        <html>
            <head></head>
            <body>
                <h3 style="margin-left:140px">TERCEIRO TURNO</h3>
                <p>Relatório gerado em: <strong>{0}</strong></p>
                <p>Segue Relatórios de todas as impressoras e estoque de suprimentos do 3º Turno:</p>
                <p></p>
                {1}
                <p></p>
                <p>Toners MFP 432FDN em estoque : <strong>{2}</strong></p>
                <p></p>
                {3}
                <p></p>
                <p>Toners E52645 em estoque : <strong>{4}</strong></p>
                <p></p>
                {5}
                <p></p>
                <p>
                Toner M479 <strong>Preto</strong> em estoque : <strong>{6}</strong></br>
                Toner M479 <strong>Ciano</strong> em estoque : <strong>{7}</strong></br>
                Toner M479 <strong>Magenta</strong> em estoque : <strong>{8}</strong></br>
                Toner M479 <strong>Amarelo</strong> em estoque : <strong>{9}</strong></br>
                </p>
                <p></p>
                <p>Abra a planilha dos Relatórios Diarios <a href="\\\srvsao028\Automação Python\Relatorio_Diario.xlsx"
                                                            target="_blank">Clicando aqui</a>
                </p>
                <p>Abra a planilha de Controle de Estoque Simpress <a href="\\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx""
                                                                    target="_blank">Clicando aqui</a>
                </p>
            <p>Suporte TI - (11) 4152-9518 / 9821</p>
            </body>
        </html>
        """.format(self.GeradoEm,self.relatorio1.to_html(),self.Toner_MFP432.to_string(index=False),self.relatorio2.to_html(),self.Toner_MFPE52645.to_string(index=False),self.relatorio3.to_html(),self.M479_Black.to_string(index=False),self.M479_Cyan.to_string(index=False),self.M479_Pink.to_string(index=False),self.M479_Yellow.to_string(index=False))
        self.body = MIMEText(self.html, 'html')
        self.msg.attach(self.body)
        self.filename = "Relatorio_Diario.xlsx"
        self.attachment = open(r"\\SRVSAO028\Automação Python\Relatorio_Diario.xlsx","rb") 
        self.p = MIMEBase('application', 'octet-stream') 
        self.p.set_payload((self.attachment).read()) 
        encoders.encode_base64(self.p) 
        self.p.add_header('Content-Disposition', "attachment; filename= %s" % self.filename) 
        self.msg.attach(self.p) 
        self.s = smtplib.SMTP('email.plural.com.br') 
        self.s.ehlo()
        #self.s.login("Sistemas.plural","asdf321!@#") 
        self.text = self.msg.as_string() 
        self.s.sendmail(self.fromaddr, self.toaddr, self.text) 
        self.s.quit() 
        print('E-mail Enviado com Sucesso!')
        time.sleep(3)


        #FORA DE TURNO
    def MFP_432 (self):
        self.offline = 0
        print('Modelo MFP 432...')
        self.lista_dos_toner = []
        self.lista_da_imagem = []
        self.nome_impressoras = ['Jurídico', 'Segurança Trabalho',
                                'Ambulatório Dr Assunção','TI',
                                'RH','Pré Impressão',
                                'Ambulatório','Comercial',
                                'Papel e Tinta','Expedição',
                                'Manutenção','Portaria',
                                'Produção','Impressão Digital',
                                'Recepção']
        self.modelo_mfp_432 = [
        'https://10.10.4.150/sws/index.html', #Juridico
        'https://10.10.4.173/sws/index.html', #Segurança do Trabalho
        'https://10.10.4.177/sws/index.html', #Ambulatório Dr Assunção
        'https://10.10.4.174/sws/index.html', #TI
        'https://10.10.4.157/sws/index.html', #RH
        'https://10.10.4.151/sws/index.html', #Pré-Impressão
        'https://10.10.4.155/sws/index.html', #Ambulatório
        'https://10.10.4.162/sws/index.html', #Comercial
        'https://10.10.4.164/sws/index.html', #Papel e Tinta
        'https://10.10.4.163/sws/index.html', #Expedição
        'https://10.10.4.158/sws/index.html', #Manutenção
        'https://10.10.4.160/sws/index.html', #Portaria
        'https://10.10.4.156/sws/index.html', #Produção
        'https://10.10.4.169/sws/index.html', #Impressão Digital
        'https://10.10.4.182/sws/index.html', #Recepção
        ]

        print('Abrindo Navegador')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())

        for self.i in range(15):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_mfp_432[self.i])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()                
                print("Procurando Dados...")
                self.toner = 1
                while(self.toner == 1):
                    try:
                        self.toner = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="ext-gen300"]/div/table/tbody/tr/td[2]/div/div/div[2]')))
                        self.lista_dos_toner.append(self.toner.text)
                        time.sleep(2)
                        print('Armazenando dados do Toner...')
                        self.imagem = self.navegador.find_element(By.XPATH,'//*[@id="ext-gen344"]/div/table/tbody/tr/td[2]/div/div/div[2]')
                        self.lista_da_imagem.append(self.imagem.text)
                        time.sleep(2)
                        print('Armazenando dados da Imagem...')    
                        print(self.lista_dos_toner, self.lista_da_imagem)  
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()   
                        self.toner = 1   
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx",sheet_name='Planilha 1')
                
                self.MFP432 = self.planilha.iloc[2:17,:4].rename(columns= {'Relatório Avulso':'Local','Unnamed: 1':'Modelo',
                                                                'Unnamed: 2':f'% de Toner','Unnamed: 3':'Uni.Imagem'})
                
                self.Toner = self.MFP432.loc[self.MFP432["Local"] == self.nome_impressoras[self.i],f"% de Toner"]
                self.Imagem = self.MFP432.loc[self.MFP432["Local"] == self.nome_impressoras[self.i],"Uni.Imagem"]

                self.lista_dos_toner.append(self.Toner.to_string(index=False))
                self.lista_da_imagem.append(self.Imagem.to_string(index=False)) 
                print(self.lista_dos_toner, self.lista_da_imagem) 

        print("Impressoras MFP 432 Offline: ",self.offline)
        print('Fechando Navegador...')
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_432 (self):
        print('Carregando Planilha...')
        self.index = 4
        self.book= openpyxl.load_workbook(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx")
        self.planilha = self.book['Planilha 1']
        self.DiaHoraAtual = datetime.now().strftime("%d/%m/%Y %H:%M")
        self.planilha.cell(column=6, row=1, value=self.DiaHoraAtual)
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.lista_dos_toner, self.lista_da_imagem):
                self.planilha.cell(column=3, row=self.index, value=self.nome)
                self.planilha.cell(column=4, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_E52645(self):
        self.offline = 0
        print('Modelo MFP E52645...')
        self.cartucho_preto = []
        self.nome_impressoras = ['ALMOXARIFADO','QUALIDADE','ATENDIMENTO','FINANCEIRO','COMPRAS']
        self.kit_aliment_documentos = []
        self.modelo_e52645 = [
        'https://10.10.4.152/hp/device/DeviceStatus/Index', #Almoxarifado
        'https://10.10.4.153/', #Qualidade
        'https://10.10.4.154/', #Atendimento PB
        'https://10.10.4.171/', #Financeiro
        'https://10.10.4.159/', #Compras
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        
        for self.l in range(5):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_e52645[self.l])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print('Procurando dados...')
                self.cartucho = 1
                while(self.cartucho == 1):
                    try:
                        self.cartucho = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="SupplyPLR0"]')))
                        print('Armazenando dados do Cartucho...')
                        self.cartuchop = self.cartucho.text.split('*')
                        self.cartucho_preto.append(self.cartuchop[0])
                        self.kit = self.navegador.find_element(By.XPATH,'//*[@id="SupplyPLR1"]')
                        print('Armazenado dados Do kit...')
                        self.kitA = self.kit.text.split('*')
                        self.kit_aliment_documentos.append(self.kitA[0])
                        time.sleep(3)
                        print(self.cartucho_preto, self.kit_aliment_documentos)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()
                        self.cartucho = 1
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx",sheet_name='Planilha 1')

                self.E52645 = self.planilha.iloc[19:24,:4].rename(columns= {'Relatório Avulso':'Local','Unnamed: 1':'Modelo',
                                                                    'Unnamed: 2':'Toner Preto','Unnamed: 3':'Kit Aliment. Doc.'})

                self.Toner = self.E52645.loc[self.E52645["Local"] == self.nome_impressoras[self.i],"Toner Preto"]   
                self.Kit = self.E52645.loc[self.E52645["Local"] == self.nome_impressoras[self.i],"Kit Aliment. Doc."]

                self.cartucho_preto.append(self.Toner.to_string(index=False))
                self.kit_aliment_documentos.append(self.Kit.to_string(index=False))
                print(self.cartucho_preto, self.kit_aliment_documentos) 
        print("Impressoras MFP E52645 Offline: ",self.offline)
        self.offline = 0
        self.navegador.quit()
    def Salvar_Dados_MFP_E52645 (self):
        print('Carregando Planilha...')
        self.index = 21
        self.book= openpyxl.load_workbook(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx")
        self.planilha = self.book['Planilha 1']
        for self.i in range(1):
            for self.nome, self.modelo in zip(self.cartucho_preto, self.kit_aliment_documentos):
                self.planilha.cell(column=3, row=self.index, value=self.nome)
                self.planilha.cell(column=4, row=self.index, value=self.modelo)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx")        
        print("Planilha Salva com Sucesso")
    def MFP_M479fdw(self):
        self.offline = 0
        print('Modelo MFP M479fdw...')
        self.nome_impressoras = ['Dir. Carlos Jacomine','Anatilia','Atendimento Color','Rogerio Cordeiro','Welinton Martins','Adriana Gasparine']
        self.cartucho_preto = []
        self.cartucho_cyan = []
        self.cartucho_magenta = []
        self.cartucho_yellow = []
        self.modelo_M479fdw = [
        'https://10.10.4.178/#hId-pgConsumables',
        'https://10.10.4.167/#hId-pgConsumables',
        'https://10.10.4.168/#hId-pgConsumables',
        'https://10.10.4.175/#hId-pgConsumables',
        'https://10.10.4.172/#hId-pgConsumables',
        'https://10.10.4.170/#hId-pgConsumables'
        ]
        print('Abrindo navegador...')
        self.navegador = webdriver.Chrome(ChromeDriverManager().install())
        

        for self.j in range(6):
            try:
                print('Carregando pagina...')
                self.navegador.get(self.modelo_M479fdw[self.j])
                self.navegador.find_element(By.XPATH,'//*[@id="details-button"]').click()
                self.navegador.find_element(By.XPATH,'//*[@id="proceed-link"]').click()
                print('Coletando dados...')
                self.preto = 1
                while(self.preto == 1):
                    try:
                        self.preto = WebDriverWait(self.navegador, 60).until(EC.presence_of_element_located((By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[2]')))
                        self.cyan = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[3]')
                        self.magenta = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[4]')
                        self.yellow = self.navegador.find_element(By.XPATH,'//*[@id="appConsumable-inkCart-tbl-Tbl"]/tbody/tr[8]/td[5]')
                        print('Armazenando dados dos Cartuchos...')
                        self.cartucho_preto.append(self.preto.text)
                        self.cartucho_cyan.append(self.cyan.text)
                        self.cartucho_magenta.append(self.magenta.text)
                        self.cartucho_yellow.append(self.yellow.text)
                        time.sleep(3)
                        print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow)
                        time.sleep(3)
                    except:
                        print("Atualizando Pagina...")
                        self.navegador.refresh()
                        self.preto = 1
            except:
                print("Error: Impressora Offline!")
                self.offline += 1
                self.planilha = pd.read_excel(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx",sheet_name='Planilha 1')

                self.MFP479 = self.planilha.iloc[26:32,:6].rename(columns= {'Relatório Avulso':'Local','Unnamed: 1':'Modelo',
                                                                    'Unnamed: 2':'Preto','Unnamed: 3':'Ciano','Unnamed: 4':'Magenta','Unnamed: 5':'Amarelo'})

                print(self.MFP479)
                self.Preto = self.MFP479.loc[self.MFP479["Local"] == self.nome_impressoras[self.j],"Preto"]                                                                 
                self.Ciano = self.MFP479.loc[self.MFP479["Local"] == self.nome_impressoras[self.j],"Ciano"]
                self.Magenta = self.MFP479.loc[self.MFP479["Local"] == self.nome_impressoras[self.j],"Magenta"]                                                                 
                self.Amarelo = self.MFP479.loc[self.MFP479["Local"] == self.nome_impressoras[self.j],"Amarelo"]
 
                self.cartucho_preto.append(self.Preto.to_string(index=False))
                self.cartucho_cyan.append(self.Ciano.to_string(index=False))
                self.cartucho_magenta.append(self.Magenta.to_string(index=False))
                self.cartucho_yellow.append(self.Amarelo.to_string(index=False))
                print(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow) 

        print("Impressoras MFP M479fdw Offline: ",self.offline)
        self.offline = 0                
        print('Fechando Navegador...')
        self.navegador.quit()    
    def Salvar_Dados_MFP_M479fdw (self):
        print('Carregando Planilha...')
        self.index = 28
        self.book= openpyxl.load_workbook(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx")
        self.planilha = self.book['Planilha 1']
        for self.i in range(1):
            for self.colorblack, self.colorcyan, self.colormagenta, self.coloryellow in zip(self.cartucho_preto, self.cartucho_cyan, self.cartucho_magenta, self.cartucho_yellow):
                self.planilha.cell(column=3, row=self.index, value=self.colorblack)
                self.planilha.cell(column=4, row=self.index, value=self.colorcyan)
                self.planilha.cell(column=5, row=self.index, value=self.colormagenta)
                self.planilha.cell(column=6, row=self.index, value=self.coloryellow)
                self.index += 1
        print('Salvando Planilha...')
        time.sleep(3)
        self.book.save(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx")        
        print("Planilha Salva com Sucesso")
    def Enviar_Email(self):
        self.tipos_colunas = {'Modelo' : str,'Estoque': str}
        self.lista_colunas = ['Modelo','Cor','Estoque']

        self.tabela_simpress = pd.read_excel(r'\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx', sheet_name='Python')

        pd.set_option('display.precision',0)

        self.Toner_MFP432 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "432FDN","Estoque"]
        self.Toner_MFPE52645 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "E52645","Estoque"]
        self.Toner_MFP7E77830 = self.tabela_simpress.loc[self.tabela_simpress["Modelo"] == "HP 7E77830","Estoque"]
        self.M479_Black = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Black","Estoque"]
        self.M479_Cyan = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Cyan","Estoque"]
        self.M479_Pink = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Pink","Estoque"]
        self.M479_Yellow = self.tabela_simpress.loc[self.tabela_simpress["Cor"] == "Yellow","Estoque"]

        print('Enviando E-mail...')
        time.sleep(3)
        
        self.planilhaAberta = pd.read_excel(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx", sheet_name='Planilha 1')
        
        
        self.relatorio1 = self.planilhaAberta.iloc[2:17,:4].rename(columns={'Relatório Avulso':'Local','Unnamed: 1':'Modelo','Unnamed: 2':f'% de Toner','Unnamed: 3':'Uni.Imagem'}).style.set_caption('Consumo de Toner HP Laser MFP 432').set_table_styles([
             { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
              'padding-right': '20px'}).hide_index()
        self.relatorio2 = self.planilhaAberta.iloc[19:24,:4].rename(columns={'Relatório Avulso':'Local','Unnamed: 1':'Modelo','Unnamed: 2':'Toner Preto','Unnamed: 3':'Kit Aliment. Doc.'}).style.set_caption('Consumo de Toner HP LaserJet MFP E52645').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()
        self.relatorio3 = self.planilhaAberta.iloc[26:32,:6].rename(columns={'Relatório Avulso':'Local','Unnamed: 1':'Modelo','Unnamed: 2':'Preto','Unnamed: 3':'Ciano','Unnamed: 4':'Magenta','Unnamed: 5':'Amarelo'}).style.set_caption('Consumo de Toner HP LaserJet Pro MFP M479fdw').set_table_styles([
            { 'selector': 'caption', 'props': 'font-size: 18px; font-weight: bold; text-align: center' }]).set_properties(**{'border':'1px solid black',
             'padding-right': '20px'}).hide_index()


        self.fromaddr = "sistemas.plural@plural.com.br"
        self.toaddr = "kelvin.rocha@plural.com.br"

        self.msg = MIMEMultipart() 
        self.msg['From'] = self.fromaddr 
        self.msg['To'] = self.toaddr 
        self.msg['Subject'] = "Relatório de Consumo e Suprimentos (Simpress)"
        self.html = """\
        <html>
            <head></head>
            <body>
                <h3 style="margin-left:140px">Relatório Avulso</h3>
                <p>Relatório gerado em: <strong>{0}</strong></p>
                <p>Segue Relatórios de todas as impressoras e estoque de suprimentos:</p>
                <p></p>
                {1}
                <p></p>
                <p>Toners MFP 432FDN em estoque : <strong>{2}</strong></p>
                <p></p>
                {3}
                <p></p>
                <p>Toners E52645 em estoque : <strong>{4}</strong></p>
                <p></p>
                {5}
                <p></p>
                <p>
                Toner M479 <strong>Preto</strong> em estoque : <strong>{6}</strong></br>
                Toner M479 <strong>Ciano</strong> em estoque : <strong>{7}</strong></br>
                Toner M479 <strong>Magenta</strong> em estoque : <strong>{8}</strong></br>
                Toner M479 <strong>Amarelo</strong> em estoque : <strong>{9}</strong></br>
                </p>
                <p></p>
                <p>Abra a planilha do Relatório Avulso  <a href="\\\srvsao028\Automação Python\Relatorio_Avulso.xlsx"
                                                            target="_blank">Clicando aqui</a>
                </p>
                <p>Abra a planilha de Controle de Estoque Simpress <a href="\\\srvsao040\Departamentos\TI\Suporte\Estoque Simpress\Estoque (Simpress).xlsx""
                                                                    target="_blank">Clicando aqui</a>
                </p>
            <p>Suporte TI - (11) 4152-9518 / 9821</p>
            </body>
        </html>
        """.format(self.GeradoEm,self.relatorio1.to_html(),self.Toner_MFP432.to_string(index=False),self.relatorio2.to_html(),self.Toner_MFPE52645.to_string(index=False),self.relatorio3.to_html(),self.M479_Black.to_string(index=False),self.M479_Cyan.to_string(index=False),self.M479_Pink.to_string(index=False),self.M479_Yellow.to_string(index=False))
        self.body = MIMEText(self.html, 'html')
        self.msg.attach(self.body)
        self.filename = "Relatorio_Avulso.xlsx"
        self.attachment = open(r"\\srvsao028\Automação Python\Relatorio_Avulso.xlsx","rb") 
        self.p = MIMEBase('application', 'octet-stream') 
        self.p.set_payload((self.attachment).read()) 
        encoders.encode_base64(self.p) 

        self.p.add_header('Content-Disposition', "attachment; filename= %s" % self.filename) 
        self.msg.attach(self.p) 
        self.s = smtplib.SMTP('email.plural.com.br') 
        self.s.ehlo()
        #self.s.login("Sistemas.plural","asdf321!@#") 
        self.text = self.msg.as_string() 
        self.s.sendmail(self.fromaddr, self.toaddr, self.text) 
        self.s.quit() 
        print('E-mail Enviado com Sucesso!')
        time.sleep(3)

Start = Relatorio_Impressoras()
Start.Inicio()